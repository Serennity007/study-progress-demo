# -*- coding: utf-8 -*-
"""成都智慧象留学 · 学生进度追踪系统 后端服务

Flask + SQLite 零配置实现，接口契约对齐《商业企划书》第六章表 6：
  POST /api/auth/login              登录并签发令牌
  POST /api/students                新建学员档案（教师）
  GET  /api/students                学员列表（教师）
  GET  /api/students/{id}           学员档案与汇总进度（教师；学员仅限本人）
  POST /api/students/{id}/records   登记学习记录（教师）
  PUT  /api/records/{id}/review     评阅记录并填写评语（教师）
  POST /api/students/{id}/scores    录入阶段测评成绩（教师）
  GET  /api/stats/overview          进度总览统计（教师）

扩展接口：
  GET  /api/directory               学号目录（登录页，无需令牌）
  GET  /api/bootstrap               当前身份可见的全量数据（页面首屏一次拉取）
  PUT  /api/me/password             修改密码
  GET  /api/subjects / POST /api/subjects   自定义科目（落库）
  GET  /api/export/students.xlsx    学员进度总览导出（Excel，支持学期筛选）
  PUT  /api/students/{id}/status    停用/启用学员（教师）
  GET  /api/config                  班级列表 + 阶段划分标准（教师）
  POST /api/classes                新增班级（教师）
  PUT  /api/classes/{name}/status   停用/启用班级（教师）
  PUT  /api/stage-standards/{name}  修改阶段划分标准说明（教师）
  GET  /api/audit                   操作审计查询（教师）
  POST /api/parent/login           家长凭学号+授权码登录
  GET  /api/parent/summary         家长查看学习摘要（仅限授权学员）

同时以 no-cache 头托管上层目录的静态页面（登录页/老师端/学员端/家长端）。
关键操作（登录、档案变更、评阅、成绩、导出、密码、配置）均写入审计日志。

启动：python backend/app.py   →  http://localhost:8686
数据：backend/zhxx.db（首次运行自动从 seed.json 建库，删除该文件即可重置）
"""
import hashlib
import json
import os
import secrets
import sqlite3
from datetime import date, datetime, timedelta

from flask import Flask, g, jsonify, request, send_file, send_from_directory

BASE_DIR = os.path.dirname(os.path.abspath(__file__))          # .../03-Demo网站/backend
STATIC_DIR = os.path.dirname(BASE_DIR)                          # .../03-Demo网站
DB_PATH = os.path.join(BASE_DIR, "zhxx.db")
SEED_PATH = os.path.join(BASE_DIR, "seed.json")

app = Flask(__name__, static_folder=None)

# 内置科目目录（与 js/common.js 的 SUBJECT_CATALOG 保持一致）
SUBJECT_CATALOG = {
    "雅思": ["雅思听力", "雅思口语", "雅思阅读", "雅思写作"],
    "托福": ["托福阅读", "托福听力", "托福口语", "托福写作"],
    "A-Level": ["A-Level 数学", "A-Level 物理", "A-Level 化学", "A-Level 经济"],
    "AP": ["AP 微积分", "AP 物理", "AP 化学", "AP 经济学", "AP 计算机科学A"],
}

STAGE_SEED = [
    ("基础期", "词汇语法与基础技能筑底，建立学习习惯与错题本体系"),
    ("强化期", "分项强化与套题训练，集中突破薄弱科目"),
    ("冲刺期", "全真模考与考前状态调整，锁定目标分"),
]


# ---------------- 数据库 ----------------

def db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exc):
    conn = g.pop("db", None)
    if conn is not None:
        conn.close()


def hash_pw(password):
    return hashlib.sha256(("zhxx:" + password).encode("utf-8")).hexdigest()


def ensure_schema():
    """幂等建表 + 旧库迁移（新增列），保证老 zhxx.db 平滑升级。"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.executescript(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            account TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,             -- teacher / student
            name TEXT NOT NULL,
            title TEXT DEFAULT '',
            student_id TEXT
        );
        CREATE TABLE IF NOT EXISTS students (
            id TEXT PRIMARY KEY,
            student_no TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            grade TEXT DEFAULT '',
            class_name TEXT DEFAULT '',
            subjects TEXT NOT NULL DEFAULT '[]',
            enroll_date TEXT DEFAULT '',
            stage TEXT DEFAULT '基础期',
            goal TEXT DEFAULT '',
            progress REAL DEFAULT 0,
            trend TEXT NOT NULL DEFAULT '[]',
            weekly_hours TEXT NOT NULL DEFAULT '[0,0,0,0,0,0,0]'
        );
        CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id TEXT NOT NULL REFERENCES students(id),
            date TEXT NOT NULL,
            subject TEXT NOT NULL,
            content TEXT NOT NULL,
            duration REAL NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT '已完成',
            comment TEXT DEFAULT '',
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS exams (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id TEXT NOT NULL REFERENCES students(id),
            stage TEXT NOT NULL,
            subject TEXT NOT NULL,
            score REAL NOT NULL,
            date TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS custom_subjects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS sessions (
            token TEXT PRIMARY KEY,
            account TEXT NOT NULL,
            role TEXT NOT NULL,
            name TEXT NOT NULL,
            title TEXT DEFAULT '',
            student_id TEXT,
            created_at TEXT NOT NULL
        );
        """
    )
    # 旧库迁移：students 增加 状态 / 家长授权码 两列
    cols = {r[1] for r in c.execute("PRAGMA table_info(students)").fetchall()}
    if "status" not in cols:
        c.execute("ALTER TABLE students ADD COLUMN status TEXT NOT NULL DEFAULT '在读'")
    if "auth_code" not in cols:
        c.execute("ALTER TABLE students ADD COLUMN auth_code TEXT NOT NULL DEFAULT ''")
    # 新业务表
    c.executescript(
        """
        CREATE TABLE IF NOT EXISTS classes (
            name TEXT PRIMARY KEY,
            active INTEGER NOT NULL DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS stage_standards (
            name TEXT PRIMARY KEY,
            description TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts TEXT NOT NULL,
            actor TEXT NOT NULL,
            action TEXT NOT NULL,
            target TEXT DEFAULT '',
            detail TEXT DEFAULT ''
        );
        """
    )
    conn.commit()
    conn.close()


def seed_if_empty():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    if c.execute("SELECT COUNT(*) FROM students").fetchone()[0] > 0:
        conn.close()
        return
    with open(SEED_PATH, "r", encoding="utf-8") as f:
        seed = json.load(f)
    pw = hash_pw(seed["credentials"]["studentPassword"])
    teacher = seed["credentials"]["teacher"]
    c.execute(
        "INSERT INTO users(account,password_hash,role,name,title) VALUES(?,?,?,?,?)",
        (teacher["account"], hash_pw(teacher["password"]), "teacher", teacher["name"], teacher["title"]),
    )
    now = datetime.now().isoformat(timespec="seconds")
    for s in seed["students"]:
        c.execute(
            "INSERT INTO users(account,password_hash,role,name,student_id) VALUES(?,?,?,?,?)",
            (s["studentNo"], pw, "student", s["name"], s["id"]),
        )
        c.execute(
            "INSERT INTO students(id,student_no,name,grade,class_name,subjects,enroll_date,stage,goal,progress,trend,weekly_hours,status,auth_code)"
            " VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                s["id"], s["studentNo"], s["name"], s["grade"], s["className"],
                json.dumps(s["subjects"], ensure_ascii=False),
                s["enrollDate"], s["stage"], s["goal"],
                s["progress"], json.dumps(s["trend"]), json.dumps(s["weeklyHours"]),
                "在读", "P" + s["studentNo"][1:],
            ),
        )
        for r in reversed(s["records"]):  # 倒序插入，保证最新记录 id 最大
            c.execute(
                "INSERT INTO records(student_id,date,subject,content,duration,status,comment,created_at)"
                " VALUES(?,?,?,?,?,?,?,?)",
                (s["id"], r["date"], r["subject"], r["content"], r["duration"], r["status"], r.get("comment", ""), now),
            )
        for e in s["exams"]:
            c.execute(
                "INSERT INTO exams(student_id,stage,subject,score,date) VALUES(?,?,?,?,?)",
                (s["id"], e["stage"], e["subject"], e["score"], e["date"]),
            )
        if s["className"]:
            c.execute("INSERT OR IGNORE INTO classes(name, active) VALUES(?,1)", (s["className"],))
    # 家长账号（参照 frappe/education 与 lav_sms 的家长实体做法：家长为独立用户，挂接学员）
    for s in seed["students"]:
        c.execute(
            "INSERT INTO users(account,password_hash,role,name,student_id) VALUES(?,?,?,?,?)",
            ("P" + s["studentNo"][1:], pw, "parent", s["name"] + "家长", s["id"]),
        )
    for name, desc in STAGE_SEED:
        c.execute("INSERT OR IGNORE INTO stage_standards(name, description) VALUES(?,?)", (name, desc))
    conn.commit()
    conn.close()
    print("已初始化数据库 zhxx.db（种子数据 10 名学员）")


def ensure_parent_accounts():
    """幂等：为每个学员补建家长账号（账号 P+学号数字，默认密码 zx123456）。"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    rows = c.execute(
        "SELECT s.id, s.student_no, s.name FROM students s "
        "WHERE NOT EXISTS (SELECT 1 FROM users u WHERE u.role='parent' AND u.student_id=s.id)"
    ).fetchall()
    for sid, sno, name in rows:
        c.execute(
            "INSERT INTO users(account,password_hash,role,name,student_id) VALUES(?,?,?,?,?)",
            ("P" + sno[1:], hash_pw("zx123456"), "parent", name + "家长", sid),
        )
    conn.commit()
    conn.close()


def init_db():
    ensure_schema()
    seed_if_empty()
    ensure_parent_accounts()


def log_action(user, action, target="", detail=""):
    db().execute(
        "INSERT INTO audit_log(ts, actor, action, target, detail) VALUES(?,?,?,?,?)",
        (datetime.now().isoformat(timespec="seconds"), user.get("name", "?"), action, target, detail),
    )


# ---------------- 序列化 ----------------

def student_row_to_dict(row, with_detail=False):
    out = {
        "id": row["id"],
        "name": row["name"],
        "studentNo": row["student_no"],
        "grade": row["grade"],
        "className": row["class_name"],
        "subjects": json.loads(row["subjects"]),
        "enrollDate": row["enroll_date"],
        "stage": row["stage"],
        "goal": row["goal"],
        "progress": row["progress"],
        "trend": json.loads(row["trend"]),
        "weeklyHours": json.loads(row["weekly_hours"]),
        "status": row["status"] if "status" in row.keys() else "在读",
        "authCode": row["auth_code"] if "auth_code" in row.keys() else "",
        "parentAccount": "P" + row["student_no"][1:],
    }
    if with_detail:
        cur = db().execute(
            "SELECT * FROM records WHERE student_id=? ORDER BY date DESC, id DESC", (row["id"],)
        )
        out["records"] = [dict(r) for r in cur.fetchall()]
        cur = db().execute(
            "SELECT * FROM exams WHERE student_id=? ORDER BY date DESC, id DESC", (row["id"],)
        )
        out["exams"] = [dict(r) for r in cur.fetchall()]
    return out


def current_user():
    header = request.headers.get("Authorization", "")
    token = header[7:] if header.startswith("Bearer ") else ""
    if not token:
        return None
    row = db().execute("SELECT * FROM sessions WHERE token=?", (token,)).fetchone()
    return dict(row) if row else None


def require_role(role):
    user = current_user()
    if not user:
        return None, (jsonify({"error": "未登录或会话已过期"}), 401)
    if role == "teacher" and user["role"] != "teacher":
        return None, (jsonify({"error": "需要教师权限"}), 403)
    return user, None


def custom_subjects():
    rows = db().execute("SELECT name FROM custom_subjects ORDER BY id").fetchall()
    return [r["name"] for r in rows]


def subjects_payload():
    groups = [{"group": k, "items": list(v)} for k, v in SUBJECT_CATALOG.items()]
    custom = custom_subjects()
    if custom:
        groups.append({"group": "自定义", "items": custom})
    return groups


def bump_progress(student_id, step):
    """推进总进度与趋势末位（与前端演示规则一致：加记录 +1，录成绩 +2，上限 99）"""
    conn = db()
    row = conn.execute("SELECT progress, trend FROM students WHERE id=?", (student_id,)).fetchone()
    progress = min(99, round((row["progress"] + step) * 10) / 10)
    trend = json.loads(row["trend"]) or [0]
    trend[-1] = progress
    conn.execute(
        "UPDATE students SET progress=?, trend=? WHERE id=?",
        (progress, json.dumps(trend), student_id),
    )
    conn.commit()


def semester_of(date_str):
    """按入学日期归入学期：2–7 月为春季学期，8–1 月为秋季学期（跨年归入学年起点）。"""
    try:
        d = date.fromisoformat(str(date_str))
    except (TypeError, ValueError):
        return "未知"
    if 2 <= d.month <= 7:
        return f"{d.year} 春季学期"
    return f"{d.year if d.month >= 8 else d.year - 1} 秋季学期"


# ---------------- 静态页面（no-cache，行为同 server.py） ----------------

@app.after_request
def no_cache(resp):
    resp.headers.setdefault("Cache-Control", "no-store, no-cache, must-revalidate")
    resp.headers.setdefault("Pragma", "no-cache")
    resp.headers.setdefault("Expires", "0")
    return resp


@app.route("/", defaults={"path": "index.html"})
@app.route("/<path:path>")
def static_files(path):
    target = os.path.normpath(os.path.join(STATIC_DIR, path))
    if not target.startswith(STATIC_DIR) or not os.path.isfile(target):
        return "Not Found", 404
    return send_from_directory(STATIC_DIR, path)


# ---------------- 认证 ----------------

@app.post("/api/ping")
def ping():
    return jsonify({"ok": True, "mode": "server"})


@app.get("/api/directory")
def directory():
    rows = db().execute(
        "SELECT student_no, name FROM students WHERE status='在读' ORDER BY student_no"
    ).fetchall()
    return jsonify({"students": [{"studentNo": r["student_no"], "name": r["name"]} for r in rows]})


@app.post("/api/auth/login")
def login():
    body = request.get_json(silent=True) or {}
    account = str(body.get("account", "")).strip()
    password = str(body.get("password", ""))
    row = db().execute("SELECT * FROM users WHERE account=?", (account,)).fetchone()
    if not row or row["password_hash"] != hash_pw(password):
        return jsonify({"error": "账号或密码不正确"}), 401
    if row["role"] in ("student", "parent"):
        st = db().execute("SELECT status, name FROM students WHERE id=?", (row["student_id"],)).fetchone()
        if st and st["status"] == "停用":
            return jsonify({"error": "该学员账号已停用，请联系老师"}), 403
    token = secrets.token_hex(16)
    conn = db()
    conn.execute(
        "INSERT INTO sessions(token,account,role,name,title,student_id,created_at) VALUES(?,?,?,?,?,?,?)",
        (token, row["account"], row["role"], row["name"], row["title"] or "", row["student_id"], datetime.now().isoformat(timespec="seconds")),
    )
    conn.commit()
    log_action({"name": row["name"]}, "登录", row["account"])
    conn.commit()
    return jsonify({
        "token": token,
        "profile": {
            "role": row["role"], "name": row["name"], "title": row["title"] or "",
            "account": row["account"], "id": row["student_id"], "studentNo": row["account"] if row["role"] == "student" else "",
        },
    })


@app.post("/api/auth/logout")
def logout():
    user = current_user()
    if user:
        db().execute("DELETE FROM sessions WHERE token=?", (user["token"],))
        db().commit()
    return jsonify({"ok": True})


@app.get("/api/bootstrap")
def bootstrap():
    user, err = require_role("any")
    if err:
        return err
    conn = db()
    if user["role"] == "teacher":
        rows = conn.execute("SELECT * FROM students ORDER BY status DESC, student_no").fetchall()
        students = [student_row_to_dict(r, with_detail=True) for r in rows]
    elif user["role"] == "parent":
        rows = conn.execute("SELECT * FROM students WHERE id=?", (user["student_id"],)).fetchall()
        students = [student_row_to_dict(r, with_detail=True) for r in rows]
    else:
        row = conn.execute("SELECT * FROM students WHERE id=?", (user["student_id"],)).fetchone()
        students = [student_row_to_dict(row, with_detail=True)] if row else []
    return jsonify({
        "profile": {"role": user["role"], "name": user["name"], "title": user["title"], "id": user["student_id"]},
        "students": students,
        "subjects": subjects_payload(),
    })


@app.put("/api/me/password")
def change_password():
    user, err = require_role("any")
    if err:
        return err
    body = request.get_json(silent=True) or {}
    old, new = str(body.get("oldPassword", "")), str(body.get("newPassword", ""))
    if len(new) < 6:
        return jsonify({"error": "新密码至少 6 位"}), 400
    row = db().execute("SELECT password_hash FROM users WHERE account=?", (user["account"],)).fetchone()
    if not row or row["password_hash"] != hash_pw(old):
        return jsonify({"error": "原密码不正确"}), 400
    db().execute("UPDATE users SET password_hash=? WHERE account=?", (hash_pw(new), user["account"]))
    log_action(user, "修改密码", user["account"])
    db().commit()
    return jsonify({"ok": True})


# ---------------- 学员与记录（表 6 契约 + 扩展） ----------------

@app.get("/api/students")
def list_students():
    user, err = require_role("teacher")
    if err:
        return err
    rows = db().execute("SELECT * FROM students ORDER BY status DESC, student_no").fetchall()
    return jsonify({"students": [student_row_to_dict(r) for r in rows]})


@app.post("/api/students")
def create_student():
    user, err = require_role("teacher")
    if err:
        return err
    body = request.get_json(silent=True) or {}
    name = str(body.get("name", "")).strip()
    if not name:
        return jsonify({"error": "姓名不能为空"}), 400
    subjects = body.get("subjects") or []
    if not subjects:
        return jsonify({"error": "请至少勾选一门备考课程"}), 400
    conn = db()
    nums = [int(r["student_no"][1:]) for r in conn.execute("SELECT student_no FROM students").fetchall()
            if r["student_no"][1:].isdigit()]
    next_no = "S" + str(max(nums) + 1) if nums else "S2026001"
    student_id = next_no
    default_pw = hash_pw("zx123456")
    c = conn.cursor()
    c.execute(
        "INSERT INTO users(account,password_hash,role,name,student_id) VALUES(?,?,?,?,?)",
        (next_no, default_pw, "student", name, student_id),
    )
    # 家长账号与学员同步建立（家长登录互不影响学生端会话）
    c.execute(
        "INSERT INTO users(account,password_hash,role,name,student_id) VALUES(?,?,?,?,?)",
        ("P" + next_no[1:], default_pw, "parent", name + "家长", student_id),
    )
    c.execute(
        "INSERT INTO students(id,student_no,name,grade,class_name,subjects,enroll_date,stage,goal,progress,trend,weekly_hours,status,auth_code)"
        " VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (
            student_id, next_no, name,
            str(body.get("grade", "高一")),
            str(body.get("className", "")).strip(),
            json.dumps(subjects, ensure_ascii=False),
            str(body.get("enrollDate", date.today().isoformat())),
            str(body.get("stage", "基础期")),
            str(body.get("goal", "")).strip(),
            0, json.dumps([0, 0, 0, 0, 0, 0]), json.dumps([0] * 7),
            "在读", "P" + next_no[1:],
        ),
    )
    log_action(user, "新建学员", next_no, f"{name}（初始密码 zx123456）")
    conn.commit()
    fresh = conn.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
    return jsonify({"student": student_row_to_dict(fresh, with_detail=True)})


@app.get("/api/students/<student_id>")
def get_student(student_id):
    user, err = require_role("any")
    if err:
        return err
    if user["role"] != "teacher" and user["student_id"] != student_id:
        return jsonify({"error": "学员仅能读取本人数据（RBAC）"}), 403
    row = db().execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
    if not row:
        return jsonify({"error": "学员不存在"}), 404
    return jsonify({"student": student_row_to_dict(row, with_detail=True)})


@app.put("/api/students/<student_id>")
def update_student(student_id):
    user, err = require_role("teacher")
    if err:
        return err
    body = request.get_json(silent=True) or {}
    row = db().execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
    if not row:
        return jsonify({"error": "学员不存在"}), 404
    name = str(body.get("name", "")).strip() or row["name"]
    subjects = body.get("subjects") or json.loads(row["subjects"])
    db().execute(
        "UPDATE students SET name=?, grade=?, class_name=?, stage=?, subjects=?, enroll_date=?, goal=? WHERE id=?",
        (
            name,
            str(body.get("grade", row["grade"])),
            str(body.get("className", row["class_name"])),
            str(body.get("stage", row["stage"])),
            json.dumps(subjects, ensure_ascii=False),
            str(body.get("enrollDate", row["enroll_date"])),
            str(body.get("goal", row["goal"])),
            student_id,
        ),
    )
    db().execute("UPDATE users SET name=? WHERE student_id=?", (name, student_id))
    log_action(user, "编辑学员档案", student_id, name)
    db().commit()
    fresh = db().execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
    return jsonify({"student": student_row_to_dict(fresh, with_detail=True)})


@app.put("/api/students/<student_id>/status")
def set_student_status(student_id):
    user, err = require_role("teacher")
    if err:
        return err
    body = request.get_json(silent=True) or {}
    status = body.get("status")
    if status not in ("在读", "停用"):
        return jsonify({"error": "状态取值不合法"}), 400
    reason = str(body.get("reason", "")).strip()
    row = db().execute("SELECT name FROM students WHERE id=?", (student_id,)).fetchone()
    if not row:
        return jsonify({"error": "学员不存在"}), 404
    db().execute("UPDATE students SET status=? WHERE id=?", (status, student_id))
    if status == "停用":
        # 停用即踢出该学员及其家长的全部会话
        db().execute("DELETE FROM sessions WHERE student_id=?", (student_id,))
    log_action(user, "停用学员" if status == "停用" else "启用学员", student_id,
               row["name"] + (("（原因：" + reason + "）") if reason else ""))
    db().commit()
    fresh = db().execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
    return jsonify({"student": student_row_to_dict(fresh, with_detail=True)})


@app.post("/api/students/<student_id>/records")
def add_record(student_id):
    user, err = require_role("teacher")
    if err:
        return err
    body = request.get_json(silent=True) or {}
    date_s = str(body.get("date", "")).strip()
    content = str(body.get("content", "")).strip()
    duration = body.get("duration")
    try:
        duration = round(float(duration), 1)
    except (TypeError, ValueError):
        return jsonify({"error": "学习时长格式不正确"}), 400
    if not date_s or not content or duration <= 0:
        return jsonify({"error": "请完整填写日期、学习内容与时长"}), 400
    subject = str(body.get("subject", "")).strip() or "未分类"
    status = body.get("status") if body.get("status") in ("已完成", "待评阅") else "已完成"
    comment = str(body.get("comment", "")).strip()

    conn = db()
    conn.execute(
        "INSERT INTO records(student_id,date,subject,content,duration,status,comment,created_at)"
        " VALUES(?,?,?,?,?,?,?,?)",
        (student_id, date_s, subject, content, duration, status, comment, datetime.now().isoformat(timespec="seconds")),
    )
    # 累计本周时长（仅统计本周的记录）
    try:
        d = date.fromisoformat(date_s)
        today = date.today()
        monday = today - timedelta(days=today.weekday())
        if monday <= d <= today:
            idx = (d - monday).days
            row = conn.execute("SELECT weekly_hours, name FROM students WHERE id=?", (student_id,)).fetchone()
            hours = json.loads(row["weekly_hours"]) or [0] * 7
            hours[idx] = round(hours[idx] + duration, 1)
            conn.execute("UPDATE students SET weekly_hours=? WHERE id=?", (json.dumps(hours), student_id))
    except ValueError:
        pass
    conn.commit()
    bump_progress(student_id, 1)
    log_action(user, "登记学习记录", student_id, f"{subject} · {content[:30]}")
    db().commit()
    fresh = conn.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
    return jsonify({"student": student_row_to_dict(fresh, with_detail=True)})


@app.put("/api/records/<int:record_id>/review")
def review_record(record_id):
    user, err = require_role("teacher")
    if err:
        return err
    body = request.get_json(silent=True) or {}
    comment = str(body.get("comment", "")).strip()
    if not comment:
        return jsonify({"error": "请填写教师评语"}), 400
    row = db().execute(
        "SELECT r.student_id AS student_id, s.name AS name FROM records r JOIN students s ON s.id=r.student_id WHERE r.id=?",
        (record_id,),
    ).fetchone()
    if not row:
        return jsonify({"error": "记录不存在"}), 404
    db().execute("UPDATE records SET comment=?, status='已完成' WHERE id=?", (comment, record_id))
    log_action(user, "评阅记录", row["student_id"], comment[:30])
    db().commit()
    fresh = db().execute("SELECT * FROM students WHERE id=?", (row["student_id"],)).fetchone()
    return jsonify({"student": student_row_to_dict(fresh, with_detail=True)})


@app.post("/api/students/<student_id>/scores")
def add_score(student_id):
    user, err = require_role("teacher")
    if err:
        return err
    body = request.get_json(silent=True) or {}
    try:
        score = round(float(body.get("score")), 1)
    except (TypeError, ValueError):
        return jsonify({"error": "分数格式不正确"}), 400
    date_s = str(body.get("date", "")).strip()
    if not date_s or score < 0:
        return jsonify({"error": "请完整填写分数与测评日期"}), 400
    subject = str(body.get("subject", "")).strip() or "未分类"
    conn = db()
    conn.execute(
        "INSERT INTO exams(student_id,stage,subject,score,date) VALUES(?,?,?,?,?)",
        (student_id, str(body.get("stage", "基础期")), subject, score, date_s),
    )
    conn.commit()
    bump_progress(student_id, 2)
    log_action(user, "录入测评成绩", student_id, f"{subject} {score}")
    db().commit()
    fresh = conn.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
    return jsonify({"student": student_row_to_dict(fresh, with_detail=True)})


@app.get("/api/stats/overview")
def stats_overview():
    user, err = require_role("teacher")
    if err:
        return err
    rows = db().execute("SELECT progress, weekly_hours FROM students WHERE status='在读'").fetchall()
    students = [student_row_to_dict(r, with_detail=True) for r in
                db().execute("SELECT * FROM students WHERE status='在读'").fetchall()]
    hours = sum(sum(json.loads(r["weekly_hours"] or "[]")) for r in rows)
    progress = sum(r["progress"] for r in rows) / len(rows) if rows else 0
    pending = sum(1 for s in students for rec in s.get("records", []) if rec["status"] == "待评阅")
    return jsonify({
        "total": len(rows),
        "weeklyHours": round(hours, 1),
        "avgProgress": round(progress),
        "pending": pending,
    })


# ---------------- 科目库（自定义科目落库） ----------------

@app.get("/api/subjects")
def get_subjects():
    return jsonify({"groups": subjects_payload()})


@app.post("/api/subjects")
def add_subject():
    user, err = require_role("teacher")
    if err:
        return err
    body = request.get_json(silent=True) or {}
    name = str(body.get("name", "")).strip()
    if not name:
        return jsonify({"error": "科目名称不能为空"}), 400
    known = any(name in g_["items"] for g_ in subjects_payload())
    if not known:
        db().execute(
            "INSERT OR IGNORE INTO custom_subjects(name, created_at) VALUES(?,?)",
            (name, datetime.now().isoformat(timespec="seconds")),
        )
        log_action(user, "新增科目", name)
        db().commit()
    return jsonify({"groups": subjects_payload()})


# ---------------- 班级与阶段标准配置（UC-A07） ----------------

@app.get("/api/config")
def get_config():
    user, err = require_role("teacher")
    if err:
        return err
    classes = [dict(r) for r in db().execute("SELECT name, active FROM classes ORDER BY name").fetchall()]
    stages = [dict(r) for r in db().execute("SELECT name, description FROM stage_standards").fetchall()]
    return jsonify({"classes": classes, "stageStandards": stages})


@app.post("/api/classes")
def add_class():
    user, err = require_role("teacher")
    if err:
        return err
    body = request.get_json(silent=True) or {}
    name = str(body.get("name", "")).strip()
    if not name:
        return jsonify({"error": "班级名称不能为空"}), 400
    exists = db().execute("SELECT 1 FROM classes WHERE name=?", (name,)).fetchone()
    if exists:
        return jsonify({"error": "该班级已存在"}), 409
    db().execute("INSERT INTO classes(name, active) VALUES(?,1)", (name,))
    log_action(user, "新增班级", name)
    db().commit()
    classes = [dict(r) for r in db().execute("SELECT name, active FROM classes ORDER BY name").fetchall()]
    return jsonify({"classes": classes})


@app.put("/api/classes/<class_name>/status")
def set_class_status(class_name):
    user, err = require_role("teacher")
    if err:
        return err
    body = request.get_json(silent=True) or {}
    active = 1 if body.get("active") else 0
    row = db().execute("SELECT active FROM classes WHERE name=?", (class_name,)).fetchone()
    if not row:
        return jsonify({"error": "班级不存在"}), 404
    db().execute("UPDATE classes SET active=? WHERE name=?", (active, class_name))
    log_action(user, "启用班级" if active else "停用班级", class_name)
    db().commit()
    classes = [dict(r) for r in db().execute("SELECT name, active FROM classes ORDER BY name").fetchall()]
    return jsonify({"classes": classes})


@app.put("/api/stage-standards/<stage_name>")
def update_stage_standard(stage_name):
    user, err = require_role("teacher")
    if err:
        return err
    body = request.get_json(silent=True) or {}
    description = str(body.get("description", "")).strip()
    row = db().execute("SELECT 1 FROM stage_standards WHERE name=?", (stage_name,)).fetchone()
    if not row:
        return jsonify({"error": "阶段不存在"}), 404
    db().execute("UPDATE stage_standards SET description=? WHERE name=?", (description, stage_name))
    log_action(user, "修改阶段标准", stage_name, description[:30])
    db().commit()
    stages = [dict(r) for r in db().execute("SELECT name, description FROM stage_standards").fetchall()]
    return jsonify({"stageStandards": stages})


# ---------------- 操作审计（UC-A08） ----------------

@app.get("/api/audit")
def audit_query():
    user, err = require_role("teacher")
    if err:
        return err
    action = request.args.get("action", "").strip()
    limit = min(int(request.args.get("limit", 100) or 100), 500)
    if action:
        rows = db().execute(
            "SELECT * FROM audit_log WHERE action=? ORDER BY id DESC LIMIT ?", (action, limit)
        ).fetchall()
    else:
        rows = db().execute(
            "SELECT * FROM audit_log ORDER BY id DESC LIMIT ?", (limit,)
        ).fetchall()
    return jsonify({"logs": [dict(r) for r in rows]})


# ---------------- Excel 导出（UC-A06：支持学期筛选） ----------------

@app.get("/api/export/students.xlsx")
def export_students():
    user, err = require_role("teacher")
    if err:
        return err
    semester = request.args.get("semester", "").strip()
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "学员进度总览"
    header_fill = PatternFill("solid", fgColor="12233A")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="微软雅黑", size=10)
    thin = Side(style="thin", color="D9D2C0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["学号", "姓名", "年级", "班级", "当前阶段", "备考课程", "入学学期", "备考进度(%)",
               "本周时长(小时)", "学习记录(条)", "待评阅(条)", "入学日期", "状态", "阶段目标"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows = db().execute("SELECT * FROM students ORDER BY status DESC, student_no").fetchall()
    exported = 0
    for r in rows:
        s = student_row_to_dict(r, with_detail=True)
        sem = semester_of(s["enrollDate"])
        if semester and semester != "全部学期" and sem != semester:
            continue
        pending = sum(1 for rec in s["records"] if rec["status"] == "待评阅")
        sem_records = sum(1 for rec in s["records"]
                          if semester_of(rec["date"]) == semester) if semester and semester != "全部学期" else len(s["records"])
        ws.append([
            s["studentNo"], s["name"], s["grade"], s["className"], s["stage"],
            "、".join(s["subjects"]), sem, round(s["progress"], 1),
            round(sum(s["weeklyHours"]), 1), len(s["records"]), pending,
            s["enrollDate"], s["status"], s["goal"],
        ])
        exported += 1
    widths = [11, 9, 7, 13, 9, 26, 13, 11, 12, 11, 10, 12, 8, 34]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    path = os.path.join(BASE_DIR, "export_tmp.xlsx")
    wb.save(path)
    log_action(user, "导出数据", semester or "全部学期", f"共 {exported} 名学员")
    db().commit()
    fname = "学员进度总览.xlsx" if not semester or semester == "全部学期" else f"学员进度总览_{semester}.xlsx"
    return send_file(
        path,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------- 家长查看通道（UC-S06） ----------------

@app.post("/api/parent/login")
def parent_login():
    body = request.get_json(silent=True) or {}
    student_no = str(body.get("studentNo", "")).strip()
    auth_code = str(body.get("authCode", "")).strip()
    row = db().execute(
        "SELECT * FROM students WHERE student_no=? AND auth_code=? AND status='在读'",
        (student_no, auth_code),
    ).fetchone()
    if not row:
        return jsonify({"error": "学号或授权码不正确"}), 401
    token = secrets.token_hex(16)
    conn = db()
    conn.execute(
        "INSERT INTO sessions(token,account,role,name,title,student_id,created_at) VALUES(?,?,?,?,?,?,?)",
        (token, student_no, "parent", f"{row['name']}家长", "", row["id"], datetime.now().isoformat(timespec="seconds")),
    )
    conn.commit()
    log_action({"name": f"{row['name']}家长"}, "家长登录", student_no)
    conn.commit()
    return jsonify({
        "token": token,
        "profile": {"role": "parent", "name": row["name"], "id": row["id"], "studentNo": row["student_no"]},
    })


@app.get("/api/parent/summary")
def parent_summary():
    user, err = require_role("any")
    if err:
        return err
    if user["role"] != "parent":
        return jsonify({"error": "仅家长身份可访问"}), 403
    row = db().execute("SELECT * FROM students WHERE id=?", (user["student_id"],)).fetchone()
    if not row:
        return jsonify({"error": "学员不存在"}), 404
    s = student_row_to_dict(row, with_detail=True)
    comments = [
        {"date": r["date"], "subject": r["subject"], "comment": r["comment"]}
        for r in s["records"] if r["comment"]
    ][:3]
    return jsonify({
        "student": {
            "name": s["name"], "studentNo": s["studentNo"], "grade": s["grade"],
            "className": s["className"], "stage": s["stage"], "goal": s["goal"],
            "progress": s["progress"], "weeklyHours": s["weeklyHours"], "subjects": s["subjects"],
        },
        "recentComments": comments,
        "latestExams": s["exams"][:3],
    })


if __name__ == "__main__":
    init_db()
    print("成都智慧象留学 · 学生进度追踪系统 后端已启动")
    print("访问 http://localhost:8686   （数据文件：backend/zhxx.db，删除即重置）")
    app.run(host="0.0.0.0", port=8686, threaded=True)
